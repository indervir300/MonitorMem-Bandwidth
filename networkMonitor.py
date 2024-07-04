import tkinter as tk
import psutil
from openpyxl import Workbook, load_workbook
from datetime import datetime
import json
import os

UPDATE_DELAY = 1  # 1 second delay

def get_size(bytes):
    """Returns size of bytes in a nice format."""
    for unit in ['', 'K', 'M', 'G', 'T', 'P']:
        if bytes < 1024:
            return f"{bytes:.2f}{unit}B"
        bytes /= 1024

def save_cumulative_stats(upload, download, last_reset_day):
    with open("cumulative_stats.json", "w") as f:
        json.dump({"upload": upload, "download": download, "last_reset_day": last_reset_day}, f)

def load_cumulative_stats():
    if os.path.exists("cumulative_stats.json"):
        with open("cumulative_stats.json", "r") as f:
            stats = json.load(f)
            return stats["upload"], stats["download"], stats["last_reset_day"]
    return 0, 0, None

# Initialize variables to store previous values
prev_cumulative_upload = 0
prev_cumulative_download = 0

def update_stats():
    global io_1, cumulative_upload, cumulative_download, last_reset_day
    global prev_cumulative_upload, prev_cumulative_download
    
    try:
        io_2 = psutil.net_io_counters()
    except psutil.Error as e:
        # Handle network interface errors gracefully, log or display a message
        print(f"Error fetching network stats: {e}")
        io_2 = None
    
    if io_2:
        bytes_sent, bytes_recv = io_2.bytes_sent, io_2.bytes_recv
        
        # Check for zeros in stats
        if bytes_sent == 0:
            cumulative_upload = prev_cumulative_upload
        else:
            cumulative_upload += bytes_sent - io_1.bytes_sent
        
        if bytes_recv == 0:
            cumulative_download = prev_cumulative_download
        else:
            cumulative_download += bytes_recv - io_1.bytes_recv
        
        # Update previous values
        prev_cumulative_upload = cumulative_upload
        prev_cumulative_download = cumulative_download
        
        # Update labels
        upload_label.config(text=f"Up: {get_size(cumulative_upload)}")
        download_label.config(text=f"Down: {get_size(cumulative_download)}")
        upload_speed_label.config(text=f"Up Speed: {get_size((bytes_sent - io_1.bytes_sent) / UPDATE_DELAY)}/s")
        download_speed_label.config(text=f"Down Speed: {get_size((bytes_recv - io_1.bytes_recv) / UPDATE_DELAY)}/s")
        
        # Check if the day has changed to reset stats
        now = datetime.now()
        current_day = now.strftime("%A")
        if current_day != last_reset_day:
            cumulative_upload = 0
            cumulative_download = 0
            last_reset_day = current_day
            save_cumulative_stats(cumulative_upload, cumulative_download, last_reset_day)
        else:
            save_cumulative_stats(cumulative_upload, cumulative_download, last_reset_day)
        
        # Update the bytes_sent and bytes_recv for the next iteration
        io_1 = io_2
    
    root.after(UPDATE_DELAY * 1000, update_stats)


def close_window():
    root.destroy()

def start_drag(event):
    root.x = event.x
    root.y = event.y

def drag_window(event):
    deltax = event.x - root.x
    deltay = event.y - root.y
    x = root.winfo_x() + deltax
    y = root.winfo_y() + deltay
    root.geometry(f"+{x}+{y}")

def save_to_excel(data):
    try:
        # Try to load existing workbook
        wb = load_workbook("network_bandwidth_report.xlsx")
    except FileNotFoundError:
        # Create a new workbook if not exists
        wb = Workbook()
        ws = wb.active
        ws.title = "Network Stats"
        ws.append(["Date", "Time", "Upload", "Download"])

    ws = wb.active
    ws.append(data)
    wb.save("network_bandwidth_report.xlsx")

def export_data():
    io = psutil.net_io_counters()
    bytes_sent, bytes_recv = io.bytes_sent, io.bytes_recv

    now = datetime.now()
    date = now.strftime("%d-%m-%Y")
    time = now.strftime("%H:%M:%S")
    upload = get_size(bytes_sent)
    download = get_size(bytes_recv)

    data = [date, time, upload, download]
    save_to_excel(data)

root = tk.Tk()
root.title("Network Bandwidth Monitor")
root.configure(bg="#212121")
root.overrideredirect(True)  # Remove default title bar

# Initialize a variable to track the "stay on top" state
stay_on_top = False

# Function to handle toggling "stay on top"
def toggle_stay_on_top():
    global stay_on_top
    stay_on_top = not stay_on_top
    root.attributes("-topmost", stay_on_top)
    stay_on_top_button.config(text="Pin" if not stay_on_top else "Unpin")

# Create a custom title bar
title_bar = tk.Frame(root, bg="#262626", pady=5)
title_bar.pack(fill="x")

title_bar.bind("<ButtonPress-1>", start_drag)
title_bar.bind("<B1-Motion>", drag_window)

# Title label
title_label = tk.Label(title_bar, text="Network Monitor", bg="#262626", fg="#f5f5f5", font=("Verdana", 8, "bold"))
title_label.pack(side="left", padx=2)

# Title bar height based on the title label's font size
title_bar_height = title_label.winfo_reqheight() + 20  # Add some padding
title_bar.config(height=title_bar_height)

# Create a close button
close_button = tk.Button(title_bar, text="X", command=close_window, bg="#ff0000", fg="#f5f5f5", font=("Arial", 9, "bold"), width=4, relief="flat")
close_button.pack(side="right", padx=4)

# Create a "Stay on Top" button
stay_on_top_button = tk.Button(title_bar, text="Pin", command=toggle_stay_on_top, bg="#212121", fg="#f5f5f5", font=("Arial", 9), width=7, relief="flat")
stay_on_top_button.pack(side="right", padx=2)

# Create a "Save or Export" button
export_button = tk.Button(title_bar, text="Save", command=export_data, bg="#212121", fg="#f5f5f5", font=("Arial", 9), width=7, relief="flat")
export_button.pack(side="right", padx=2)

# Initialize the initial network stats
io_1 = psutil.net_io_counters()
cumulative_upload, cumulative_download, last_reset_day = load_cumulative_stats()

if last_reset_day is None:
    last_reset_day = datetime.now().strftime("%A")

# Create a main frame
main_frame = tk.Frame(root, bg="#212121")
main_frame.pack(fill="both", expand=True, pady=10, padx=5)

# Create labels
upload_label = tk.Label(main_frame, text=f"Up: {get_size(cumulative_upload)}", font=("Arial", 9, "bold"), fg="#f5f5f5", bg="#212121")
download_label = tk.Label(main_frame, text=f"Down: {get_size(cumulative_download)}", font=("Arial", 9, "bold"), fg="#f5f5f5", bg="#212121")
upload_speed_label = tk.Label(main_frame, text="Up Speed: N/A", font=("Arial", 9, "bold"), fg="#f5f5f5", bg="#212121")
download_speed_label = tk.Label(main_frame, text="Down Speed: N/A", font=("Arial", 9, "bold"), fg="#f5f5f5", bg="#212121")

upload_label.grid(row=0, column=0, padx=5, sticky="ew", pady=5)
download_label.grid(row=0, column=1, padx=5, sticky="ew", pady=5)
upload_speed_label.grid(row=0, column=2, padx=5, sticky="ew", pady=5)
download_speed_label.grid(row=0, column=3, padx=5, sticky="ew", pady=5)


main_frame.grid_columnconfigure(0, weight=1)
main_frame.grid_columnconfigure(1, weight=1)
main_frame.grid_columnconfigure(2, weight=1)
main_frame.grid_columnconfigure(3, weight=1)


root.update_idletasks()
width = main_frame.winfo_reqwidth() + 150
height = main_frame.winfo_reqheight() + title_bar_height + 15
root.geometry(f"{width}x{height}")

root.after(0, update_stats)

root.mainloop()
